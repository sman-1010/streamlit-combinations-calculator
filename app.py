import streamlit as st
import pandas as pd
from collections import Counter
from io import BytesIO

# Only import openpyxl if we do color-coding
# (We will conditionally import it later if needed)

def parse_list(input_str):
    """Helper to parse a comma-separated string of integers."""
    arr = []
    for x in input_str.split(','):
        x = x.strip()
        if x.lstrip('-').isdigit():  # handle negative if needed
            arr.append(int(x))
    return arr

def remove_nwis(sample_list, nwis):
    """Remove any items from sample_list that appear in nwis."""
    return [x for x in sample_list if x not in nwis]

def is_valid_triple(A, B, C, counts, strict_switch, nwis):
    """
    Check if the triple (A, B, C) can be formed with available counts.
    If strict_switch is True, apply extra checks involving nwis.
    """
    from collections import Counter
    temp_counts = Counter([A, B, C])
    
    # If strict_switch is on, apply extra NWIS logic
    if strict_switch:
        if (B - 5) in nwis:
            return False
        if (C - B + 5) in nwis:
            return False

    # Check we have enough occurrences for each number
    for num, req in temp_counts.items():
        if counts.get(num, 0) < req:
            return False
    return True

def is_valid_double(A, B, counts, strict_switch, nwis):
    """
    Check if the double (A, B) can be formed with available counts.
    If strict_switch is True, apply extra checks involving nwis.
    """
    from collections import Counter
    temp_counts = Counter([A, B])
    
    # If strict_switch is on, apply extra NWIS logic
    if strict_switch:
        if (B - 1) in nwis:
            return False

    # Check we have enough occurrences for each number
    for num, req in temp_counts.items():
        if counts.get(num, 0) < req:
            return False
    return True

def compute_bins(triple, Main, G, R, C_list):
    """
    Calculate how many numbers in the triple come from
    each bin: [Main, G, R, C_list].
    """
    bins = [0, 0, 0, 0]  # [M_count, G_count, R_count, C_count]

    # Copies so original data is not modified
    main_copy = Main.copy()
    g_copy    = G.copy()
    r_copy    = R.copy()
    c_copy    = C_list.copy()

    for num in triple:
        if num in main_copy:
            bins[0] += 1
            main_copy.remove(num)
        elif num in g_copy:
            bins[1] += 1
            g_copy.remove(num)
        elif num in r_copy:
            bins[2] += 1
            r_copy.remove(num)
        elif num in c_copy:
            bins[3] += 1
            c_copy.remove(num)
    return bins

def main():
    st.title("Number Combinations Generator")

    # ---------------------------------------------------------------
    # 1) CONFIGURATION FLAGS - user selectable
    # ---------------------------------------------------------------

    # Alternative to segmented_control
    # Bold and large title with no spacing before radio buttons
    st.write("### **Select One of the Combinations Below:**")  # This avoids extra spacing

    # Radio button with hidden label to remove extra space
    method_selections = st.radio("Options:", ["Combination 1: Triples", "Combination 2: Doubles"], label_visibility="collapsed")

    # st.write(f"You selected: {method_selection}")


    # method_selection = ''
    # if method_selections == "Combination 1: Triples":
    #     # st.write("You selected Combination 1: Triples")
    #     st.write("##### You selected Combination 1: Triples")
    #     st.write("""
    #     Three numbers (B,C,SUM) selected from Main,G,R,C with conditions
    #     - SUM = C + 5
    #     - Intermediate Values of B and C are greater than 0
    #     """)

    #     method_selection = 'triple'

    # elif method_selections == "Combination 2: Doubles":
    #     st.write("##### You selected Combination 2: Doubles")
    #     st.write("""
    #     Two numbers (B,SUM) selected from Main,G,R,C with conditions
    #     - SUM = B + 4
    #     - Intermediate Value of B is greater than 0
    #     """)
    #     method_selection = 'double'

    method_selection = ''
    if method_selections == "Combination 1: Triples":
        method_selection = 'triple'
        description = """
        <b>Three numbers (B, C, SUM) selected from Main, G, R, C with conditions:</b>
        <ul>
            <li>SUM = C + 5</li>
            <li>Intermediate Values of B and C are greater than 0</li>
        </ul>
        """
    elif method_selections == "Combination 2: Doubles":
        method_selection = 'double'
        description = """
        <b>Two numbers (B, SUM) selected from Main, G, R, C with conditions:</b>
        <ul>
            <li>SUM = B + 4</li>
            <li>Intermediate Value of B is greater than 0</li>
        </ul>
        """

    # Custom CSS for grey rectangle with rounded edges
    st.markdown(
        f"""
        <div style="
            background-color: #f0f0f0; 
            padding: 15px; 
            border-radius: 10px; 
            border: 1px solid #d3d3d3;
            box-shadow: 2px 2px 5px rgba(0,0,0,0.1);
            font-size: 16px;">
            <h4 style="color: #333;">You selected {method_selections}</h4>
            {description}
        </div>
        """,
        unsafe_allow_html=True
    )




    st.markdown("---")
    # st.subheader("Select Files to save below:")
    st.subheader("Not Wanted in List:")
    # save_valid = st.checkbox("Save valid combinations (valid_combinations_final.xlsx)", value=True)
    save_valid = True
    # save_no_color = st.checkbox("Save no-color version (visualized_no_color.xlsx)", value=True)
    save_no_color = False
    # save_with_color = st.checkbox("Save color version (visualized_with_color.xlsx)", value=True)
    save_with_color = True

    # Toggle buttons for user selection
    toggle_B_C = st.toggle("Strict B and C", value=False)
    strict_switch = st.toggle("Strict Intermediate Values", value=False)
    toggle_sum = st.toggle("Strict Sum", value=False)    

    st.markdown("---")
    st.subheader("Input Configuration")

    # ---------------------------------------------------------------
    # 2) DATA AND PARAMETERS - user editable
    # ---------------------------------------------------------------

    # strict_switch = st.checkbox("Enable strict_switch (NWIS check in intermediate steps)", value=False)
    default_main = "1,3,5,9,11,13,15,16,21,23,24,25,29,31,32,33,35,37,39,41,45,47,48,52,57,65,67,68,82"
    main_str = st.text_area("Main list", default_main, height=80)

    default_g = "3,5,6,11,13,15,16,24,31,32,35"
    g_str = st.text_area("G list", default_g, height=80)

    default_r = "4,10,12,14,15,16,20,22,23,24,28,32,33,34,41"
    r_str = st.text_area("R list", default_r, height=80)

    default_c_list = "12,14,22,32"
    c_list_str = st.text_area("C_list", default_c_list, height=80)

    default_nwim = "2,4,9,10,12,14,19,20,26,27,28,34,36,42,43,44,46,49,50,53,54,56,58,59,60,62,64,66,69,70,72,73,74,76,78,79,80,21,23,26,28,29,33,39,22,4,10,12,22,28,34,4,9,10,14,19,20,28,34,44,9,10,14,19,20,22,28,30,34,40,44,50,53,54,56,59,60,70,80"
    nwim_str = st.text_area("Not Wanted List", default_nwim, height=100)
    if st.button("Run Combinations Logic"):
        # Parse user inputs
        not_wanted_in_sum = parse_list(nwim_str)
        nwis = list(set(not_wanted_in_sum))

        Main   = parse_list(main_str)
        G      = parse_list(g_str)
        R      = parse_list(r_str)
        C_list = parse_list(c_list_str)

        # ---------------------------------------------------------------
        # 4) MAIN LOGIC
        # ---------------------------------------------------------------
        # 4a) Filter out NWIS items
        # Main   = remove_nwis(Main, nwis)
        # G      = remove_nwis(G, nwis)
        # R      = remove_nwis(R, nwis)
        # C_list = remove_nwis(C_list, nwis)

        # 4b) Create a major list and get counts
        major_list = Main + G + R + C_list
        counts = Counter(major_list)
        unique_sorted = sorted(counts.keys())
        if method_selection == 'triple':
            # 4c) Generate valid triples    
            valid_triples = []
            for C in unique_sorted:
                if toggle_B_C and C in nwis:
                    continue
                A = C + 5  # Condition from the original code
                if A not in counts:
                    continue
                if toggle_sum and A in nwis:
                    continue
                for B in unique_sorted:
                    if toggle_B_C and B in nwis:
                        continue                        
                    if 5 < B < A:
                        if is_valid_triple(A, B, C, counts, strict_switch, nwis):
                            valid_triples.append((A, B, C))
            
            # 4d) Sort the triples and build a dataframe
            triple_bins = [
                (triple, compute_bins(triple, Main, G, R, C_list)) 
                for triple in valid_triples
            ]
            triple_bins_sorted = sorted(triple_bins, key=lambda x: x[0][0])

            rows = []
            for triple, bins_count in triple_bins_sorted:
                A, B, C_ = triple
                # Row: [B, C, A] + [M_count, G_count, R_count, C_count]
                rows.append([B, C_, A] + bins_count)

            columns = ['B', 'C', 'SUM', 'M_count', 'G_count', 'R_count', 'C_count']
            df = pd.DataFrame(rows, columns=columns)

            # Keep only rows that used exactly 3 items
            df = df[df[['M_count', 'G_count', 'R_count', 'C_count']].sum(axis=1) == 3]
            
            # With intermediate values
            rows = []
            for triple, bins_count in triple_bins_sorted:
                A, B, C_ = triple

                # Row: [B, C, A] + [M_count, G_count, R_count, C_count]
                rows.append([B, C_, A]+[B-5, C_-B+5] + bins_count)

            columns_triple = ['B', 'C', 'SUM','Inter 1','Inter 2', 'M_count', 'G_count', 'R_count', 'C_count']
            df_with_inter = pd.DataFrame(rows, columns=columns_triple)

            # Keep only rows that used exactly 3 items
            df_with_inter = df_with_inter[df_with_inter[['M_count', 'G_count', 'R_count', 'C_count']].sum(axis=1) == 3]
            
            st.success("Combinations generated!")

            st.write(f"Number of valid rows: {len(df_with_inter)}")
            df_with_inter.index = range(1, len(df_with_inter) + 1)
            st.dataframe(df_with_inter)  # Show a sample

            # Creating the new DataFrame with empty spaces
            new_rows = []
            for _, row in df_with_inter.iterrows():
                new_rows.append(['', row['B'], row['C'], '', '', 'Main', 'G', 'R', 'C'])
                new_rows.append([5, row['Inter 1'], row['Inter 2'], row['SUM'], '', row['M_count'], row['G_count'], row['R_count'], row['C_count']])
                new_rows.append([''] * 9)  # Empty row

            # Creating the new DataFrame
            columns_triple = ['-', 'B', 'C','SUM','--', 'M_count', 'G_count', 'R_count', 'C_count']
            df_formatted = pd.DataFrame(new_rows, columns=columns_triple)
            df_formatted.index = range(1, len(df_formatted) + 1)
            st.write("Combinations Visualized:")
            st.dataframe(df_formatted)  # Show a sample

            
        else:
            # 4c) Generate valid triples
            valid_doubles = []
            for B in unique_sorted:
                if toggle_B_C and B in nwis:
                    continue                
                A = B + 4  # Condition from the original code
                if A not in counts:
                    continue
                if toggle_sum and A in nwis:
                    continue                
                if B > 1:
                    if is_valid_double(A, B, counts, strict_switch, nwis):
                        valid_doubles.append((A, B))                        

        
            double_bins = [
                (double, compute_bins(double, Main, G, R, C_list)) 
                for double in valid_doubles
            ]
            double_bins_sorted = sorted(double_bins, key=lambda x: x[0][0])

            rows = []
            for double, bins_count in double_bins_sorted:
                A, B = double
                # Row: [B, C, A] + [M_count, G_count, R_count, C_count]
                rows.append([B, '', A] + bins_count)

            columns = ['B', '', 'SUM', 'M_count', 'G_count', 'R_count', 'C_count']
            df = pd.DataFrame(rows, columns=columns)
            # Keep only rows that used exactly 3 items
            df = df[df[['M_count', 'G_count', 'R_count', 'C_count']].sum(axis=1) == 2]
            
            rows = []
            for double, bins_count in double_bins_sorted:
                A, B = double
                # Row: [B, C, A] + [M_count, G_count, R_count, C_count]
                rows.append([B, '', A,B-1] + bins_count)

            columns = ['B', '', 'SUM','Inter 1', 'M_count', 'G_count', 'R_count', 'C_count']
            df_with_inter = pd.DataFrame(rows, columns=columns)
            # Keep only rows that used exactly 3 items
            df_with_inter = df_with_inter[df_with_inter[['M_count', 'G_count', 'R_count', 'C_count']].sum(axis=1) == 2]
            

            st.success("Combinations generated Double!")
            st.write(f"Number of valid rows: {len(df_with_inter)}")
            df_with_inter.index = range(1, len(df_with_inter) + 1)
            st.dataframe(df_with_inter)  # Show a sample

            # Creating the new DataFrame with empty spaces
            new_rows = []
            for _, row in df_with_inter.iterrows():
                new_rows.append(['', row['B'], '', '', '', 'Main', 'G', 'R', 'C'])
                new_rows.append([5, row['Inter 1'], '', row['SUM'], '', row['M_count'], row['G_count'], row['R_count'], row['C_count']])
                new_rows.append([''] * 9)  # Empty row

            # Creating the new DataFrame
            columns_triple = ['-', 'B', '','SUM','--', 'M_count', 'G_count', 'R_count', 'C_count']
            df_formatted = pd.DataFrame(new_rows, columns=columns_triple)
            df_formatted.index = range(1, len(df_formatted) + 1)
            st.write("Combinations Visualized:")
            st.dataframe(df_formatted)  # Show a sample


            # ---------------------------------------------------------------
            # 5) SAVE FIRST FILE: valid_combinations_final.xlsx (in memory)
            # ---------------------------------------------------------------
        if save_valid:
            valid_buffer = BytesIO()
            # Adjust file name for strict if needed
            valid_filename = "valid_combinations.xlsx"
            if strict_switch:
                valid_filename = "valid_combinations_strict.xlsx"      
            valid_filename = f"{method_selection}_{valid_filename}"      
            df_with_inter.to_excel(valid_buffer, index=False)
            valid_buffer.seek(0)
            st.download_button(
                label=f"Download {valid_filename}",
                data=valid_buffer,
                file_name=valid_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        # ---------------------------------------------------------------
        # 7) SAVE THIRD FILE: visualized_with_color.xlsx (in memory)
        # ---------------------------------------------------------------
        if save_with_color:
            # We'll build the colored Excel directly with openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill

            wb = Workbook()
            ws = wb.active

            # Define color fills
            green_fill  = PatternFill(start_color="9AE79C", fill_type="solid")
            blue_fill   = PatternFill(start_color="BBE3F1", fill_type="solid")
            yellow_fill = PatternFill(start_color="FCFFC6", fill_type="solid")

            # We want to replicate the same 3-row block structure
            idx = 0  # to iterate over df rows in groups of 1 (each row = one triple in df)
            row_idx = 1
            if method_selection == 'triple':
                while idx < len(df):
                    B_val = df.iloc[idx]['B']
                    C_val = df.iloc[idx]['C']
                    SUM_val = df.iloc[idx]['SUM']

                    # The top row
                    ws.cell(row=row_idx, column=1, value="")
                    ws.cell(row=row_idx, column=2, value="")
                    ws.cell(row=row_idx, column=3, value=B_val).fill = green_fill
                    ws.cell(row=row_idx, column=4, value=C_val).fill = blue_fill
                    ws.cell(row=row_idx, column=5, value="")
                    ws.cell(row=row_idx, column=6, value="")
                    ws.cell(row=row_idx, column=7, value="Main")
                    ws.cell(row=row_idx, column=8, value="G")
                    ws.cell(row=row_idx, column=9, value="R")
                    ws.cell(row=row_idx, column=10, value="C")

                    # The second row
                    triple_str = f"({B_val}, {C_val}, {SUM_val})"
                    row_2 = row_idx + 1
                    ws.cell(row=row_2, column=1, value=triple_str)
                    ws.cell(row=row_2, column=2, value=5)
                    ws.cell(row=row_2, column=3, value=(B_val - 5))
                    ws.cell(row=row_2, column=4, value=(C_val - (B_val - 5)))
                    ws.cell(row=row_2, column=5, value=SUM_val).fill = yellow_fill

                    # We also want to show counts in columns 7..10:
                    ws.cell(row=row_2, column=7, value=df.iloc[idx]['M_count'])
                    ws.cell(row=row_2, column=8, value=df.iloc[idx]['G_count'])
                    ws.cell(row=row_2, column=9, value=df.iloc[idx]['R_count'])
                    ws.cell(row=row_2, column=10, value=df.iloc[idx]['C_count'])

                    # The third row is blank
                    for c in range(1, 11):
                        ws.cell(row=row_idx+2, column=c, value="")

                    # Move to the next triple
                    row_idx += 3
                    idx += 1
            else:
                while idx < len(df):
                    B_val = df.iloc[idx]['B']
                    SUM_val = df.iloc[idx]['SUM']

                    # The top row
                    ws.cell(row=row_idx, column=1, value="")
                    ws.cell(row=row_idx, column=2, value="")
                    ws.cell(row=row_idx, column=3, value=B_val).fill = green_fill
                    ws.cell(row=row_idx, column=4, value="")
                    ws.cell(row=row_idx, column=5, value="")
                    ws.cell(row=row_idx, column=6, value="")
                    ws.cell(row=row_idx, column=7, value="Main")
                    ws.cell(row=row_idx, column=8, value="G")
                    ws.cell(row=row_idx, column=9, value="R")
                    ws.cell(row=row_idx, column=10, value="C")

                    # The second row
                    triple_str = f"({B_val}, {SUM_val})"
                    row_2 = row_idx + 1
                    ws.cell(row=row_2, column=1, value=triple_str)
                    ws.cell(row=row_2, column=2, value=5)
                    ws.cell(row=row_2, column=3, value=(B_val - 1))
                    # ws.cell(row=row_2, column=4, value=(C_val - (B_val - 5)))
                    ws.cell(row=row_2, column=4, value="")
                    ws.cell(row=row_2, column=5, value=SUM_val).fill = yellow_fill

                    # We also want to show counts in columns 7..10:
                    ws.cell(row=row_2, column=7, value=df.iloc[idx]['M_count'])
                    ws.cell(row=row_2, column=8, value=df.iloc[idx]['G_count'])
                    ws.cell(row=row_2, column=9, value=df.iloc[idx]['R_count'])
                    ws.cell(row=row_2, column=10, value=df.iloc[idx]['C_count'])

                    # The third row is blank
                    for c in range(1, 11):
                        ws.cell(row=row_idx+2, column=c, value="")

                    # Move to the next triple
                    row_idx += 3
                    idx += 1

            color_buffer = BytesIO()
            color_filename = "visualized_with_color.xlsx"
            if strict_switch:
                color_filename = "visualized_with_color_strict.xlsx"
            color_filename = f"{method_selection}_{color_filename}"
            wb.save(color_buffer)
            color_buffer.seek(0)

            st.download_button(
                label=f"Download {color_filename}",
                data=color_buffer,
                file_name=color_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        # st.success("All done!")

if __name__ == "__main__":
    main()
